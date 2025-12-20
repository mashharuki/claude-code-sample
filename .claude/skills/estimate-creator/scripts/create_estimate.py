#!/usr/bin/env python3
"""
御見積書作成スクリプト

原価と粗利率から販売単価を計算し、Excel形式の見積書を生成します。
"""

import json
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def calculate_selling_price(cost: float, gross_margin_rate: float) -> int:
    """
    原価と粗利率から販売単価を計算

    Args:
        cost: 原価(円)
        gross_margin_rate: 粗利率(%)

    Returns:
        販売単価(円、四捨五入した整数)

    Formula:
        販売単価 = 原価 ÷ (1 - 粗利率/100)
    """
    if gross_margin_rate <= 0 or gross_margin_rate >= 100:
        raise ValueError(f"粗利率は0%より大きく100%未満である必要があります: {gross_margin_rate}%")

    margin_rate = gross_margin_rate / 100
    selling_price = cost / (1 - margin_rate)
    return round(selling_price)


def format_currency(amount: int) -> str:
    """金額をカンマ区切りの文字列に変換"""
    return f"¥{amount:,}"


def create_estimate_excel(
    output_path: str,
    date: str,
    client_name: str,
    subject: str,
    payment_terms: str,
    items: list,
    gross_margin_rate: float
):
    """
    御見積書をExcelファイルとして生成

    Args:
        output_path: 出力ファイルパス
        date: 発行日(YYYYMMDD形式)
        client_name: クライアント名
        subject: 件名
        payment_terms: 支払条件
        items: 明細項目のリスト
            [
                {
                    "name": "項目名",
                    "quantity": 数量,
                    "unit": "単位",
                    "unit_cost": 原価(円)
                },
                ...
            ]
        gross_margin_rate: 粗利率(%)
    """

    # ワークブック作成
    wb = Workbook()
    ws = wb.active
    ws.title = "御見積書"

    # 列幅設定
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18

    # スタイル定義
    title_font = Font(name='メイリオ', size=16, bold=True)
    header_font = Font(name='メイリオ', size=11, bold=True)
    normal_font = Font(name='メイリオ', size=10)

    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # タイトル
    ws.merge_cells('B2:F2')
    ws['B2'] = '御見積書'
    ws['B2'].font = title_font
    ws['B2'].alignment = center_alignment

    # 日付をフォーマット(YYYYMMDD → YYYY年MM月DD日)
    try:
        date_obj = datetime.strptime(date, '%Y%m%d')
        formatted_date = date_obj.strftime('%Y年%m月%d日')
    except ValueError:
        formatted_date = date  # フォーマット失敗時はそのまま使用

    # 基本情報
    current_row = 4
    ws[f'B{current_row}'] = f'発行日: {formatted_date}'
    ws[f'B{current_row}'].font = normal_font

    current_row += 1
    ws[f'B{current_row}'] = f'{client_name} 御中'
    ws[f'B{current_row}'].font = Font(name='メイリオ', size=12, bold=True)

    current_row += 2
    ws[f'B{current_row}'] = f'件名: {subject}'
    ws[f'B{current_row}'].font = normal_font

    current_row += 1
    ws[f'B{current_row}'] = f'支払条件: {payment_terms}'
    ws[f'B{current_row}'].font = normal_font

    # 明細ヘッダー
    current_row += 2
    headers = ['No.', '項目', '数量', '単位', '単価', '金額']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    # 明細行
    current_row += 1
    start_detail_row = current_row
    total_amount = 0

    for idx, item in enumerate(items, start=1):
        # 販売単価を計算
        unit_cost = item['unit_cost']
        selling_price = calculate_selling_price(unit_cost, gross_margin_rate)
        quantity = item['quantity']
        amount = selling_price * quantity
        total_amount += amount

        # No.
        cell = ws.cell(row=current_row, column=1)
        cell.value = idx
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.border = thin_border

        # 項目名
        cell = ws.cell(row=current_row, column=2)
        cell.value = item['name']
        cell.font = normal_font
        cell.border = thin_border

        # 数量
        cell = ws.cell(row=current_row, column=3)
        cell.value = quantity
        cell.font = normal_font
        cell.alignment = right_alignment
        cell.border = thin_border

        # 単位
        cell = ws.cell(row=current_row, column=4)
        cell.value = item['unit']
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.border = thin_border

        # 単価(販売単価)
        cell = ws.cell(row=current_row, column=5)
        cell.value = selling_price
        cell.font = normal_font
        cell.alignment = right_alignment
        cell.number_format = '#,##0'
        cell.border = thin_border

        # 金額
        cell = ws.cell(row=current_row, column=6)
        cell.value = amount
        cell.font = normal_font
        cell.alignment = right_alignment
        cell.number_format = '#,##0'
        cell.border = thin_border

        current_row += 1

    # 合計行
    current_row += 1
    ws.merge_cells(f'B{current_row}:E{current_row}')
    cell = ws.cell(row=current_row, column=2)
    cell.value = '合計金額(税抜)'
    cell.font = header_font
    cell.alignment = right_alignment
    cell.border = thin_border
    cell.fill = header_fill

    cell = ws.cell(row=current_row, column=6)
    cell.value = total_amount
    cell.font = header_font
    cell.alignment = right_alignment
    cell.number_format = '#,##0'
    cell.border = thin_border
    cell.fill = header_fill

    # 消費税(10%)
    current_row += 1
    tax_amount = round(total_amount * 0.1)
    ws.merge_cells(f'B{current_row}:E{current_row}')
    cell = ws.cell(row=current_row, column=2)
    cell.value = '消費税(10%)'
    cell.font = normal_font
    cell.alignment = right_alignment
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=6)
    cell.value = tax_amount
    cell.font = normal_font
    cell.alignment = right_alignment
    cell.number_format = '#,##0'
    cell.border = thin_border

    # 総合計
    current_row += 1
    grand_total = total_amount + tax_amount
    ws.merge_cells(f'B{current_row}:E{current_row}')
    cell = ws.cell(row=current_row, column=2)
    cell.value = '総合計(税込)'
    cell.font = Font(name='メイリオ', size=12, bold=True)
    cell.alignment = right_alignment
    cell.border = thin_border
    cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

    cell = ws.cell(row=current_row, column=6)
    cell.value = grand_total
    cell.font = Font(name='メイリオ', size=12, bold=True)
    cell.alignment = right_alignment
    cell.number_format = '#,##0'
    cell.border = thin_border
    cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

    # 備考
    current_row += 3
    ws[f'B{current_row}'] = '【備考】'
    ws[f'B{current_row}'].font = header_font
    current_row += 1
    ws[f'B{current_row}'] = '・上記金額にて御見積申し上げます。'
    ws[f'B{current_row}'].font = normal_font
    current_row += 1
    ws[f'B{current_row}'] = '・ご不明な点がございましたら、お気軽にお問い合わせください。'
    ws[f'B{current_row}'].font = normal_font

    # ファイル保存
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_file))

    return {
        'success': True,
        'output_path': str(output_file),
        'total_amount': total_amount,
        'tax_amount': tax_amount,
        'grand_total': grand_total,
        'items_count': len(items)
    }


def main():
    """メイン処理"""
    if len(sys.argv) != 2:
        print("Usage: python create_estimate.py <json_file>", file=sys.stderr)
        sys.exit(1)

    # JSONファイルから設定を読み込み
    json_file = sys.argv[1]
    with open(json_file, 'r', encoding='utf-8') as f:
        config = json.load(f)

    try:
        result = create_estimate_excel(
            output_path=config['output_path'],
            date=config['date'],
            client_name=config['client_name'],
            subject=config['subject'],
            payment_terms=config['payment_terms'],
            items=config['items'],
            gross_margin_rate=config['gross_margin_rate']
        )

        # 結果を出力
        print(json.dumps(result, ensure_ascii=False, indent=2))

    except Exception as e:
        error_result = {
            'success': False,
            'error': str(e)
        }
        print(json.dumps(error_result, ensure_ascii=False, indent=2), file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
